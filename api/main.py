from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
import os
import time

app = FastAPI(title="RTD Backend", version="2.0.0")

# CORS para permitir acesso do Lovable ou outro frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Modelo esperado no corpo da requisição
class TickerPayload(BaseModel):
    ticker: str

# Caminho da planilha RTD
EXCEL_PATH = r"D:\Python\Sistema\RTD\RTD-python.xlsx"

@app.post("/ingest")
def ingest(payload: TickerPayload):
    """Recebe o ticker, grava no Excel (A2), espera o RTD atualizar e retorna os valores."""
    ticker = payload.ticker.strip().upper()

    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Arquivo RTD-python.xlsx não encontrado.")

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # Grava o ticker na célula A2
        ws["A2"] = ticker
        wb.save(EXCEL_PATH)

        # Aguarda alguns segundos para o RTD atualizar
        time.sleep(3)

        # Recarrega planilha com dados atualizados
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active

        data = {
            "ticker": ws["A2"].value,
            "cotacao": ws["B2"].value,
            "strike": ws["C2"].value,
            "vencimento": ws["D2"].value,
            "bid": ws["E2"].value,
            "ask": ws["F2"].value,
            "delta": ws["G2"].value,
            "theta": ws["H2"].value,
            "vol_imp": ws["I2"].value,
            "vl_ex": ws["J2"].value,
            "negocios": ws["K2"].value,
        }

        return {"message": "Ticker atualizado com sucesso!", "data": data}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar: {str(e)}")


@app.get("/health")
def health():
    """Verifica se o servidor está ativo"""
    return {"status": "ok"}


@app.get("/tickers")
def list_tickers():
    """Retorna o ticker atual no Excel"""
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Arquivo não encontrado.")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    return {"ticker": ws["A2"].value or ""}


@app.get("/latest")
def get_latest():
    """Retorna os últimos valores da linha 2"""
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Arquivo não encontrado.")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    return {
        "ticker": ws["A2"].value,
        "cotacao": ws["B2"].value,
        "strike": ws["C2"].value,
        "vencimento": ws["D2"].value,
        "bid": ws["E2"].value,
        "ask": ws["F2"].value,
        "delta": ws["G2"].value,
        "theta": ws["H2"].value,
        "vol_imp": ws["I2"].value,
        "vl_ex": ws["J2"].value,
        "negocios": ws["K2"].value,
    }
